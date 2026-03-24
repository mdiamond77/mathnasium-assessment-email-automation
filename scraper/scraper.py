"""
Mathnasium Nightly Radius Scraper (Python)

1. Logs into Radius
2. Downloads the Student Report (all enrolled students)
3. Scrapes student page URLs from the report table
4. Parses report to identify who needs an email draft
5. Downloads Learning Plan PDF for each triggered student
6. Pushes everything to Google Drive
"""

import os
import json
import re
import time
from datetime import datetime, timedelta
from pathlib import Path

from playwright.sync_api import sync_playwright
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

RADIUS_BASE_URL      = 'https://radius.mathnasium.com'
DOWNLOAD_DIR         = Path(__file__).parent / 'downloads'
DEBUG_DIR            = Path(__file__).parent / 'debug'
DRY_RUN              = '--dry-run' in os.sys.argv
LEVEL_UP_WINDOW_DAYS = 7

yesterday       = datetime.now() - timedelta(days=1)
yesterday_str   = f"{yesterday.month}/{yesterday.day}/{yesterday.year}"
folder_date_str = datetime.now().strftime('%Y-%m-%d')

DOWNLOAD_DIR.mkdir(exist_ok=True)
DEBUG_DIR.mkdir(exist_ok=True)


def main():
    print(f"\n🏫 Mathnasium Email Queue Builder")
    print(f"📅 Running for: {yesterday_str}")
    if DRY_RUN:
        print("🧪 DRY RUN MODE\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=['--no-sandbox', '--disable-setuid-sandbox']
        )
        context = browser.new_context(
            accept_downloads=True,
            viewport={'width': 1280, 'height': 900}
        )
        page = context.new_page()

        try:
            login(page)
            student_report_path, student_url_lookup = download_student_report(page)
            triggered_students = parse_student_report(student_report_path)

            if not triggered_students:
                print('\n✅ No students triggered today — nothing to do')
                return

            print(f"\n📋 {len(triggered_students)} student(s) need drafts today:")
            for s in triggered_students:
                print(f"  • {s['studentName']} [{s['emailType']}] — {', '.join(s['emails'])}")

            download_learning_plans(page, triggered_students, student_url_lookup)

            drive_key = os.environ.get('GOOGLE_SERVICE_ACCOUNT_KEY', '').strip()
            if not DRY_RUN and drive_key:
                upload_to_google_drive(student_report_path, triggered_students)
            elif not drive_key:
                print('\n⚠️  Skipping Drive upload — GOOGLE_SERVICE_ACCOUNT_KEY not set yet')
            else:
                print('\n🧪 DRY RUN: skipping Drive upload')

            print('\n✅ Done')

        except Exception as e:
            print(f'\n❌ Scraper failed: {e}')
            page.screenshot(path=str(DEBUG_DIR / f'error-{int(datetime.now().timestamp())}.png'))
            raise
        finally:
            browser.close()


# ─────────────────────────────────────────────
# 1. LOGIN
# ─────────────────────────────────────────────
def login(page):
    print('🔐 Logging in...')
    page.goto(RADIUS_BASE_URL)
    page.wait_for_load_state('networkidle')
    page.fill('#UserName', os.environ['RADIUS_USERNAME'])
    page.fill('#Password', os.environ['RADIUS_PASSWORD'])
    page.click('#login')
    page.wait_for_load_state('networkidle')
    try:
        page.wait_for_selector('a:has-text("Sign Out")', timeout=15000)
        print('  ✓ Logged in')
    except Exception:
        page.screenshot(path=str(DEBUG_DIR / 'login-failed.png'))
        raise Exception('Login failed — see debug/login-failed.png')


# ─────────────────────────────────────────────
# 2. DOWNLOAD STUDENT REPORT + SCRAPE URLS
# ─────────────────────────────────────────────
def download_student_report(page):
    print('\n📋 Downloading Student Report...')

    page.goto(f"{RADIUS_BASE_URL}/StudentReport")
    page.wait_for_load_state('networkidle')
    time.sleep(2)

    # Set Enrollment Filter to "Enrolled"
    page.wait_for_selector('[aria-owns="enrollmentFiltersDropDownList_listbox"]', timeout=10000)
    page.click('[aria-owns="enrollmentFiltersDropDownList_listbox"]')
    page.wait_for_selector('#enrollmentFiltersDropDownList_listbox', timeout=10000)
    time.sleep(1)
    page.click('#enrollmentFiltersDropDownList_listbox li:text-is("Enrolled")')
    time.sleep(1)

    # Search
    page.click('#btnsearch')
    page.wait_for_load_state('networkidle')
    time.sleep(10)

    # Scrape student name → Details URL from the table BEFORE exporting
    # This gives us the correct internal student ID for each student
    # Scrape student name → URL across all pages of results
    student_url_lookup = {}

    while True:
        links = page.locator('a[href*="/Student/Details/"]').all()
        # Table has separate first/last name columns — links alternate first, last, first, last
        i = 0
        while i < len(links) - 1:
            first_text = links[i].inner_text().strip().lower()
            last_text  = links[i+1].inner_text().strip().lower()
            href       = links[i].get_attribute('href')
            if href and first_text and last_text:
                student_url_lookup[f'{first_text} {last_text}'] = f'{RADIUS_BASE_URL}{href}'
            i += 2

        # Check if there is a Next page button that is enabled
        next_btn = page.locator('a.k-pager-nav[title="Go to the next page"]:not(.k-state-disabled)')
        if next_btn.count() > 0:
            next_btn.click()
            page.wait_for_load_state('networkidle')
            time.sleep(2)
        else:
            break

    print(f'  ✓ Found URLs for {len(student_url_lookup)} students')
    for k, v in list(student_url_lookup.items())[:3]:
        print(f'    {repr(k)} → {v}')

    # Screenshot for debugging
    page.screenshot(path=str(DOWNLOAD_DIR / 'before-export.png'))

    # Export to Excel
    with page.expect_download(timeout=30000) as download_info:
        page.click('#btnExport')
    download = download_info.value

    file_path = DOWNLOAD_DIR / f'student-report-{folder_date_str}.xlsx'
    download.save_as(str(file_path))
    print(f'  ✓ Saved: {file_path.name}')
    return file_path, student_url_lookup


# ─────────────────────────────────────────────
# 3. PARSE STUDENT REPORT
# ─────────────────────────────────────────────
def parse_student_report(file_path):
    print('\n🔍 Parsing Student Report...')

    wb = openpyxl.load_workbook(str(file_path))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    def get(row, col_name):
        if col_name in headers:
            val = row[headers.index(col_name)]
            return val.value if hasattr(val, 'value') else val
        return None

    triggered_students = []
    level_up_names = set()
    rows = list(ws.iter_rows(min_row=2))

    # First pass — Level Ups (Last Assessment = yesterday)
    for row in rows:
        if get(row, 'Enrollment Status') != 'Enrolled':
            continue
        last_assessment = normalize_date(get(row, 'Last Assessment'))
        if last_assessment != yesterday_str:
            continue
        name = str(get(row, 'Student Name') or '').strip()
        if not name or name == 'x x':
            continue
        emails = parse_emails(get(row, 'Guardian Emails') or get(row, 'Guardian Email List'))
        if not emails:
            print(f'  ⚠️  {name} has no guardian email — skipping')
            continue
        triggered_students.append(build_student(row, headers, get, name, emails, 'level-up', last_assessment, normalize_date(get(row, 'Last Progress Check'))))
        level_up_names.add(name.lower())
        print(f'  ✓ Level Up: {name}')

    # Second pass — Progress Checks
    for row in rows:
        if get(row, 'Enrollment Status') != 'Enrolled':
            continue
        last_pc = normalize_date(get(row, 'Last Progress Check'))
        if last_pc != yesterday_str:
            continue
        name = str(get(row, 'Student Name') or '').strip()
        if not name or name == 'x x' or name.lower() in level_up_names:
            continue
        last_assessment = normalize_date(get(row, 'Last Assessment'))
        # Within level-up window = treat as level-up (late-graded PC)
        if last_assessment:
            assess_date = parse_date(last_assessment)
            pc_date = parse_date(last_pc)
            if assess_date and pc_date:
                diff = abs((assess_date - pc_date).days)
                if diff <= LEVEL_UP_WINDOW_DAYS:
                    emails = parse_emails(get(row, 'Guardian Emails') or get(row, 'Guardian Email List'))
                    if not emails:
                        continue
                    print(f'  ✓ Level Up (late-graded PC): {name} — {diff} day(s) gap')
                    triggered_students.append(build_student(row, headers, get, name, emails, 'level-up', last_assessment, last_pc))
                    continue
        emails = parse_emails(get(row, 'Guardian Emails') or get(row, 'Guardian Email List'))
        if not emails:
            print(f'  ⚠️  {name} has no guardian email — skipping')
            continue
        triggered_students.append(build_student(row, headers, get, name, emails, 'progress-check', last_assessment, last_pc))
        print(f'  ✓ Progress Check: {name}')

    level_ups = sum(1 for s in triggered_students if s['emailType'] == 'level-up')
    pcs = sum(1 for s in triggered_students if s['emailType'] == 'progress-check')
    print(f'\n  Total: {len(triggered_students)} ({level_ups} level-up, {pcs} progress-check)')
    return triggered_students


def build_student(row, headers, get, name, emails, email_type, last_assessment, last_pc):
    return {
        'studentName':        name,
        'firstName':          name.split(' ')[0],
        'emailType':          email_type,
        'emails':             emails,
        'grade':              str(get(row, 'Grade') or ''),
        'center':             str(get(row, 'Center') or ''),
        'leadId':             str(get(row, 'Lead Id') or ''),
        'lastAssessment':     last_assessment,
        'lastProgressCheck':  last_pc,
        'learningPlanStatus': 'pending',
        'learningPlanPath':   None,
    }


# ─────────────────────────────────────────────
# 4. DOWNLOAD LEARNING PLANS
# ─────────────────────────────────────────────
def download_learning_plans(page, triggered_students, student_url_lookup):
    print('\n📚 Downloading Learning Plans...')
    for student in triggered_students:
        try:
            file_path = download_single_learning_plan(page, student, student_url_lookup)
            student['learningPlanStatus'] = 'ready'
            student['learningPlanPath'] = str(file_path)
            print(f"  ✓ {student['studentName']}")
        except Exception as e:
            print(f"  ⚠️  {student['studentName']} — LP failed: {e}")
            student['learningPlanStatus'] = 'error'
            safe = re.sub(r'[^a-z0-9]', '-', student['studentName'].lower())
            page.screenshot(path=str(DEBUG_DIR / f'lp-error-{safe}.png'))


def download_single_learning_plan(page, student, student_url_lookup):
    # Look up the correct student Details URL from the table we scraped
    name_key = student['studentName'].lower()
    url = student_url_lookup.get(name_key)

    # Fallback: try matching by first name if full name not found
    # (handles trailing spaces like "Leo " in the HTML)
    if not url:
        first = student['firstName'].lower()
        matches = {k: v for k, v in student_url_lookup.items() if k.startswith(first + ' ')}
        if len(matches) == 1:
            url = list(matches.values())[0]
            print(f"    (matched '{name_key}' via first name fallback)")

    if not url:
        raise Exception(f"No Radius URL found for '{student['studentName']}' — checked {len(student_url_lookup)} entries")

    page.goto(url)
    page.wait_for_load_state('networkidle')

    lp_locator = page.locator('a.k-grid-LPReport')
    if lp_locator.count() == 0:
        # Save screenshot to debug
        safe = re.sub(r'[^a-z0-9]', '-', student['studentName'].lower())
        page.screenshot(path=str(DOWNLOAD_DIR / f'lp-page-{safe}.png'))
        raise Exception('LP Report button not found — screenshot saved')

    with page.expect_download(timeout=15000) as dl_info:
        lp_locator.first().click()
    download = dl_info.value

    safe_name = re.sub(r'[^a-z0-9]', '-', student['studentName'].lower())
    file_path = DOWNLOAD_DIR / f"lp-{safe_name}-{folder_date_str}.pdf"
    download.save_as(str(file_path))
    return file_path


# ─────────────────────────────────────────────
# 5. UPLOAD TO GOOGLE DRIVE
# ─────────────────────────────────────────────
def upload_to_google_drive(student_report_path, triggered_students):
    print('\n☁️  Uploading to Google Drive...')

    key_data = json.loads(os.environ['GOOGLE_SERVICE_ACCOUNT_KEY'])
    creds = service_account.Credentials.from_service_account_info(
        key_data, scopes=['https://www.googleapis.com/auth/drive.file']
    )
    drive = build('drive', 'v3', credentials=creds)

    folder = drive.files().create(body={
        'name': folder_date_str,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [os.environ['GOOGLE_DRIVE_FOLDER_ID']]
    }, fields='id').execute()
    folder_id = folder['id']
    print(f'  ✓ Created folder: {folder_date_str}')

    upload_file(drive, student_report_path, folder_id,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    for student in triggered_students:
        if student['learningPlanPath']:
            upload_file(drive, Path(student['learningPlanPath']), folder_id, 'application/pdf')

    manifest = {
        'date': folder_date_str,
        'generatedAt': datetime.now().isoformat(),
        'studentReportFile': student_report_path.name,
        'triggeredStudents': [{
            'studentName':        s['studentName'],
            'firstName':          s['firstName'],
            'emailType':          s['emailType'],
            'emails':             s['emails'],
            'grade':              s['grade'],
            'center':             s['center'],
            'lastAssessment':     s['lastAssessment'],
            'lastProgressCheck':  s['lastProgressCheck'],
            'learningPlanStatus': s['learningPlanStatus'],
            'learningPlanFile':   Path(s['learningPlanPath']).name if s['learningPlanPath'] else None,
        } for s in triggered_students]
    }

    manifest_path = DOWNLOAD_DIR / f'manifest-{folder_date_str}.json'
    manifest_path.write_text(json.dumps(manifest, indent=2))
    upload_file(drive, manifest_path, folder_id, 'application/json')
    print(f"  ✓ Manifest uploaded with {len(triggered_students)} students")


def upload_file(drive, file_path, folder_id, mime_type):
    file_path = Path(file_path)
    drive.files().create(
        body={'name': file_path.name, 'parents': [folder_id]},
        media_body=MediaFileUpload(str(file_path), mimetype=mime_type)
    ).execute()
    print(f'  ✓ Uploaded: {file_path.name}')


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def parse_emails(email_str):
    if not email_str:
        return []
    return [e.strip().lower() for e in str(email_str).split(',')
            if '@' in e and 'mathnasium.com' not in e]


def normalize_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return f"{val.month}/{val.day}/{val.year}"
    s = str(val).strip().split(' ')[0]
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        return f"{int(m.group(1))}/{int(m.group(2))}/{m.group(3)}"
    return s


def parse_date(date_str):
    try:
        parts = date_str.split('/')
        return datetime(int(parts[2]), int(parts[0]), int(parts[1]))
    except Exception:
        return None


if __name__ == '__main__':
    main()
